import os, base64, mimetypes, io, json, traceback, requests
from fastapi import FastAPI, status, HTTPException
from email.message import EmailMessage
from dotenv import load_dotenv
from collections import defaultdict
from typing import List, Dict, Any
from datetime import datetime
from google.cloud import storage
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import PyPDF2
import google.generativeai as genai

load_dotenv()

USER_TOKEN_FILE = "token.json"
SERVICE_ACCOUNT_FILE = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
SENDER_USER_ID = "kevin.gianecchine@capitalexpress.cl"
EXCEL_SERVICE_URL = os.getenv("EXCEL_SERVICE_URL")
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")

print(f"DEBUG: EXCEL_SERVICE_URL = {EXCEL_SERVICE_URL}")
print(f"DEBUG: GEMINI_API_KEY = {GEMINI_API_KEY[:10]}..." if GEMINI_API_KEY else "DEBUG: GEMINI_API_KEY = None")
SCOPES = [
    "https://www.googleapis.com/auth/gmail.send",
    "https://www.googleapis.com/auth/devstorage.read_only",
]
FIXED_CC_LIST = [
    "kevin.gianecchine@capitalexpress.cl",
]
RUC_GLORIA = [
    "20100190797",
    "20600679164",
    "20312372895",
    "20524088739",
    "20467539842",
    "20506475288",
    "20418453177",
    "20512613218",
    "20115039262",
    "20100814162",
    "20518410858",
    "20101927904",
    "20479079006",
    "20100223555",
    "20532559147",
    "20487268870",
    "20562613545",
    "20478963719",
    "20481694907",
    "20454629516",
    "20512415840",
    "20602903193",
    "20392965191",
    "20601225639",
    "20547999691",
    "20600180631",
    "20116225779",
    "20131823020",
    "20601226015",
    "20131867744",
    "20603778180",
    "20131835621",
    "20511866210",
    "20481640483",
]

app = FastAPI(title="Gmail Service (HTTP Direct)")

if GEMINI_API_KEY:
    genai.configure(api_key=GEMINI_API_KEY)

def extract_pdf_data_local(pdf_paths: List[str], storage_client) -> Dict[str, Dict[str, str]]:
    """Extrae datos de PDFs usando PyPDF2 y crea diccionario para Gemini"""
    pdf_data = {}
    
    for pdf_path in pdf_paths:
        try:
            bucket_name, blob_name = pdf_path.replace("gs://", "").split("/", 1)
            pdf_bytes = storage_client.bucket(bucket_name).blob(blob_name).download_as_bytes()
            
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
            full_text = ""
            
            for page_num in range(min(2, len(pdf_reader.pages))): # Limitar a las primeras 2 páginas
                full_text += pdf_reader.pages[page_num].extract_text()
            
            clean_text = ' '.join(full_text.split())
            
            pdf_data[pdf_path] = {
                "filename": os.path.basename(pdf_path),
                "text_content": clean_text[:2000],
                "text_length": len(clean_text)
            }
            
        except Exception as e:
            print(f"Error extrayendo datos de {pdf_path}: {e}")
            pdf_data[pdf_path] = {
                "filename": os.path.basename(pdf_path),
                "text_content": "",
                "error": str(e)
            }
    
    return pdf_data

def classify_pdfs_with_gemini(pdf_data: Dict, invoices_by_ruc: Dict) -> Dict[str, List[str]]:
    """Usa Gemini para clasificar PDFs basándose en texto extraído"""
    
    if not GEMINI_API_KEY:
        print("WARN: No GEMINI_API_KEY configurado, usando fallback")
        fallback_result = defaultdict(list)
        for ruc in invoices_by_ruc.keys():
            fallback_result[ruc] = list(pdf_data.keys())
        return dict(fallback_result)
    
    known_clients = {}
    for ruc, facturas in invoices_by_ruc.items():
        known_clients[ruc] = {
            "debtor_name": facturas[0]['debtor_name'],
            "document_ids": [f['document_id'] for f in facturas],
            "client_name": facturas[0]['client_name'],
            "client_ruc": facturas[0]['client_ruc']
        }
    
    prompt = f"""
    Analiza los siguientes datos extraídos de PDFs de facturas y clasifícalos por RUC de deudor.

    CLIENTES CONOCIDOS EN ESTA OPERACIÓN:
    {json.dumps(known_clients, indent=2, ensure_ascii=False)}

    DATOS DE PDFs A CLASIFICAR:
    {json.dumps(pdf_data, indent=2, ensure_ascii=False)}

    INSTRUCCIONES:
    1. Para cada PDF, identifica el RUC del deudor/proveedor basándote en el texto extraído
    2. Coincide con uno de los RUCs conocidos en la operación
    3. Si no encuentras coincidencia exacta, usa el nombre de la empresa para hacer matching
    4. Si hay dudas, asigna a "unknown"

    RESPONDE EN ESTE FORMATO JSON:
    {{
        "classifications": {{
            "gs://path/to/pdf1.pdf": "20603596294",
            "gs://path/to/pdf2.pdf": "20100190797",
            "gs://path/to/pdf3.pdf": "unknown"
        }},
        "confidence_notes": {{
            "gs://path/to/pdf1.pdf": "Coincidencia exacta por RUC",
            "gs://path/to/pdf2.pdf": "Coincidencia por nombre empresa"
        }}
    }}
    """

    try:
        response = genai.GenerativeModel('gemini-1.5-flash').generate_content(prompt)
        print(f"DEBUG Gemini response text: {response.text}")
        
        # Limpiar respuesta de markdown si viene con ```json
        clean_text = response.text.strip()
        if clean_text.startswith('```json'):
            clean_text = clean_text[7:]  # Remove ```json
        if clean_text.endswith('```'):
            clean_text = clean_text[:-3]  # Remove ```
        clean_text = clean_text.strip()
        
        result = json.loads(clean_text)
        
        pdf_classification = defaultdict(list)
        
        for pdf_path, assigned_ruc in result["classifications"].items():
            if assigned_ruc != "unknown" and assigned_ruc in invoices_by_ruc:
                pdf_classification[assigned_ruc].append(pdf_path)
            else:
                for ruc in invoices_by_ruc.keys():
                    pdf_classification[ruc].append(pdf_path)
        
        print("Gemini classification notes:", result.get("confidence_notes", {}))
        return dict(pdf_classification)
        
    except Exception as e:
        print(f"Error en clasificación Gemini: {e}")
        fallback_result = defaultdict(list)
        for ruc in invoices_by_ruc.keys():
            fallback_result[ruc] = list(pdf_data.keys())
        return dict(fallback_result)

def get_user_credentials():
    """Obtiene las credenciales de usuario desde el archivo token.json."""
    creds = None
    if os.path.exists(USER_TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(USER_TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            from google.auth.transport.requests import Request as GoogleRequest

            creds.refresh(GoogleRequest())
        else:
            raise Exception(
                "No se encontraron credenciales de usuario válidas o el token ha expirado y no se puede refrescar."
            )
    return creds

def create_gloria_excel(invoice_data_list: List[Dict]) -> (str, bytes): # type: ignore
    """Genera un archivo Excel para Gloria a partir de una lista de diccionarios de facturas."""
    if not invoice_data_list:
        return None, None

    data_rows = []
    for invoice in invoice_data_list:
        row = {
            "FACTOR": 20603596294,
            "FECHA DE ENVIO": pd.to_datetime("today").strftime("%d/%m/%Y"),
            "RUC PROVEEDOR": invoice.get("debtor_ruc"),
            "PROVEEDOR": invoice.get("debtor_name"),
            "RUC CLIENTE": invoice.get("client_ruc"),
            "CLIENTE": invoice.get("client_name"),
            "FECHA DE EMISION": pd.to_datetime(invoice.get("issue_date"), errors="coerce").strftime("%d/%m/%Y"),
            "NUM FACTURA": invoice.get("document_id"),
            "IMPORTE NETO PAGAR": invoice.get("net_amount"),
            "MONEDA": invoice.get("currency"),
            "FECHA DE VENCIMIENTO": pd.to_datetime(invoice.get("due_date"), errors="coerce").strftime("%d/%m/%Y"),
        }
        data_rows.append(row)

    df = pd.DataFrame(data_rows)
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Facturas", index=False)
        worksheet = writer.sheets["Facturas"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        thin_border_side = Side(border_style="thin", color="000000")
        cell_border = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side,
        )
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")

        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = cell_border
                cell.alignment = center_alignment

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        for col_idx, column_cells in enumerate(worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column_letter].width = adjusted_width

        for cell in worksheet["I"][1:]:
            cell.number_format = "#,##0.00"
            cell.alignment = right_alignment

    excel_bytes = output_buffer.getvalue()
    filename = f"CapitalExpress_{datetime.now().strftime('%d%m%Y')}.xlsx"
    return filename, excel_bytes

def create_html_body(invoice_data_list: List[Dict], operation_id) -> str:
    """Crea el cuerpo HTML del correo a partir de una lista de diccionarios de facturas."""
    first_invoice = invoice_data_list[0]
    client_name = first_invoice.get("client_name")
    client_ruc = first_invoice.get("client_ruc")

    df = pd.DataFrame(invoice_data_list)
    df["total_amount"] = df.apply(
        lambda row: f"{row.get('currency', '')} {float(row.get('total_amount', 0)):,.2f}".strip(),
        axis=1,
    )
    df["net_amount"] = df.apply(
        lambda row: f"{row.get('currency', '')} {float(row.get('net_amount', 0)):,.2f}".strip(),
        axis=1,
    )
    df["due_date"] = pd.to_datetime(df["due_date"], errors="coerce").dt.strftime(
        "%d/%m/%Y"
    )

    df_display = df.rename(
        columns={
            "debtor_ruc": "RUC Deudor",
            "debtor_name": "Nombre Deudor",
            "document_id": "Documento",
            "total_amount": "Monto Factura",
            "net_amount": "Monto Neto",
            "due_date": "Fecha de Pago",
        }
    )
    display_columns = [
        "RUC Deudor",
        "Nombre Deudor",
        "Documento",
        "Monto Factura",
        "Monto Neto",
        "Fecha de Pago",
    ]
    tabla_html = df_display[display_columns].to_html(
        index=False, border=1, justify="left", classes="invoice_table"
    )

    mensaje_html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: Arial, Helvetica, sans-serif; font-size: 13px; color: #000; }}
        .container {{ max-width: 800px; }} p, li {{ line-height: 1.5; }} ol {{ padding-left: 30px; }}
        table.invoice_table {{ width: 100%; border-collapse: collapse; margin-top: 15px; margin-bottom: 20px; }}
        table.invoice_table th, table.invoice_table td {{ border: 1px solid #777; padding: 6px; text-align: left; font-size: 12px; }}
        table.invoice_table th {{ background-color: #f0f0f0; font-weight: bold; }}
        .disclaimer {{ font-style: italic; font-size: 11px; margin-top: 25px; }}
    </style>
    </head>
    <body>
    <div class="container">
        <p>Estimados señores,</p>
        <p>Por medio de la presente, les informamos que los señores de <strong>{client_name}</strong>, nos han transferido la(s) siguiente(s) factura(s) negociable(s). Solicitamos su amable confirmación sobre los siguientes puntos:</p>
        <ol>
            <li>¿La(s) factura(s) ha(n) sido recepcionada(s) conforme con sus productos o servicios?</li>
            <li>¿Cuál es la fecha programada para el pago de la(s) misma(s)?</li>
            <li>Por favor, confirmar el Monto Neto a pagar, considerando detracciones, retenciones u otros descuentos.</li>
        </ol>
        <p><strong>Detalle de las facturas:</strong></p>
        <p>Cliente: {client_name}<br>
        RUC Cliente: {client_ruc}
        </p>
        {tabla_html}
        <p>Agradecemos de antemano su pronta respuesta. Con su confirmación, procederemos a la anotación en cuenta en CAVALI.</p>
        <p class="disclaimer">"Sin perjuicio de lo anteriormente mencionado, nos permitimos recordarles que toda acción tendiente a 
        simular la emisión de la referida factura negociable o letra para obtener un beneficio a título personal o a favor de la otra 
        parte de la relación comercial, teniendo pleno conocimiento de que la misma no proviene de una relación comercial verdadera, 
        se encuentra sancionada penalmente como delito de estafa en nuestro ordenamiento jurídico.
        Asimismo, en caso de que vuestra representada cometa un delito de forma conjunta y/o en contubernio con el emitente de la factura, 
        dicha acción podría tipificarse como delito de asociación ilícita para delinquir, según el artículo 317 del Código Penal, por lo 
        que nos reservamos el derecho de iniciar las acciones penales correspondientes en caso resulte necesario"</p>
        <p>{operation_id}</p>
    </div>
    </body>
    </html>
    """
    return mensaje_html

@app.post("/send-email", status_code=status.HTTP_200_OK)
async def send_email_handler(payload: Dict[str, Any]):
    operation_id = payload.get("operation_id")
    if not operation_id:
        raise HTTPException(status_code=400, detail="Falta 'operation_id' en el payload")

    print(f"GMAIL: Recibida solicitud HTTP para op {operation_id}")

    try:
        user_creds = get_user_credentials()
        storage_client_user = storage.Client(credentials=user_creds)
        gmail_service = build('gmail', 'v1', credentials=user_creds)

        invoices = payload.get("parsed_results", [])
        if not invoices:
            print(f"WARN: No hay facturas en el payload para op {operation_id}.")
            return {"status": "skipped", "reason": "No invoices in payload"}

        invoices_by_debtor_ruc = defaultdict(list)
        for inv in invoices:
            invoices_by_debtor_ruc[inv['debtor_ruc']].append(inv)

        pdf_paths = payload.get("gcs_paths", {}).get("pdf", [])
        pdf_by_ruc = {}
        if pdf_paths:
            print(f"Clasificando {len(pdf_paths)} PDFs usando PyPDF2 + Gemini...")
            
            pdf_data = extract_pdf_data_local(pdf_paths, storage_client_user)
            
            pdf_by_ruc = classify_pdfs_with_gemini(pdf_data, invoices_by_debtor_ruc)
            
            print(f"Clasificación completada: {pdf_by_ruc}")

        sent_emails_log = []
        for ruc_deudor, facturas_grupo in invoices_by_debtor_ruc.items():
            
            correos_finales_str = ""
            try:
                nombre_deudor = facturas_grupo[0]['debtor_name']
                correo_operacion_str = payload.get('metadata', {}).get('mailVerificacion', '').strip()
                correos_operacion = [c.strip() for c in correo_operacion_str.split(';') if c.strip()]

                for correo in correos_operacion:
                    update_payload = {"ruc": ruc_deudor, "correo": correo, "nombre_deudor": nombre_deudor}
                    requests.post(f"{EXCEL_SERVICE_URL}/update-contact", json=update_payload, timeout=20).raise_for_status()

                response = requests.get(f"{EXCEL_SERVICE_URL}/get-emails/{ruc_deudor}", timeout=20)
                if response.status_code == 200:
                    correos_finales_str = response.json().get("emails", "")
                
            except requests.exceptions.RequestException as e:
                print(f"ADVERTENCIA: Falló la comunicación con Excel para RUC {ruc_deudor}. Error: {e}")
            
            if not correos_finales_str:
                print(f"ADVERTENCIA: No se encontraron correos para RUC {ruc_deudor} en op {operation_id}.")
                continue
            
            # Crear el cuerpo del mensaje HTML y el mensaje de correo
            message = EmailMessage()
            message.add_alternative(create_html_body(facturas_grupo, operation_id), subtype='html')
            message['To'] = correos_finales_str
            cc_list = set(FIXED_CC_LIST)
            if payload.get("user_email"):
                cc_list.add(payload.get("user_email"))
            message['Cc'] = ",".join(sorted(list(cc_list)))
            message['Subject'] = f"Confirmación de Facturas Negociables - {facturas_grupo[0]['client_name']}"

            # ADJUNTAR SOLO PDFs RELEVANTES PARA ESTE RUC
            relevant_pdfs = pdf_by_ruc.get(ruc_deudor, [])
            print(f"Adjuntando {len(relevant_pdfs)} PDFs para RUC {ruc_deudor}")
            for path in relevant_pdfs:
                try:
                    bucket_name, blob_name = path.replace("gs://", "").split("/", 1)
                    pdf_bytes = storage_client_user.bucket(bucket_name).blob(blob_name).download_as_bytes()
                    maintype, subtype = (mimetypes.guess_type(os.path.basename(path))[0] or "application/octet-stream").split('/')
                    message.add_attachment(pdf_bytes, maintype=maintype, subtype=subtype, filename=os.path.basename(path))
                    print(f"  ✓ Adjuntado: {os.path.basename(path)}")
                except Exception as e:
                    print(f"ADVERTENCIA: al adjuntar el archivo {path}: {e}")
                    
            # Adjuntar Excel para Gloria si corresponde
            if ruc_deudor in RUC_GLORIA:
                filename, excel_bytes = create_gloria_excel(facturas_grupo)
                if filename and excel_bytes:
                    message.add_attachment(excel_bytes, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename=filename)

            encoded_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
            gmail_service.users().messages().send(userId=SENDER_USER_ID, body={'raw': encoded_message}).execute()
            
            log_msg = f"Correo para deudor {ruc_deudor} (Op: {operation_id}) enviado a: {correos_finales_str}"
            print(f"GMAIL: {log_msg}")
            sent_emails_log.append(log_msg)

        return {"status": "success", "details": sent_emails_log}

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error interno en gmail-service: {str(e)}")
